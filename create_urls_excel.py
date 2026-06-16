from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# === ABA 1: URLs por Equipamento ===
ws1 = wb.active
ws1.title = "URLs por Equipamento"

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='1a3a6b')
brand_fills = {
    'LUMENIS': PatternFill('solid', fgColor='2d2d2d'),
    'CONTOURLINE': PatternFill('solid', fgColor='0a4a6b'),
    'CONTOURLINE MED': PatternFill('solid', fgColor='1565c0'),
    'BODY HEALTH': PatternFill('solid', fgColor='8b4a52'),
    'VISBODY': PatternFill('solid', fgColor='006b6b'),
}
brand_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
data_font = Font(name='Arial', size=10)
link_font = Font(name='Arial', size=9, color='0563C1', underline='single')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

headers = ['Marca', 'Equipamento', 'utm_brand', 'utm_term', 'URL Meta Ads (com dinamicos)', 'URL Simples (teste)']
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 95
ws1.column_dimensions['F'].width = 80
ws1.row_dimensions[1].height = 30

base = 'https://form.contourline.com.br/formulario-whatsapp.html'

equipamentos = [
    ('LUMENIS', 'TriLift', 'lumenis', 'trilift'),
    ('LUMENIS', 'UltraPulse Alpha', 'lumenis', 'ultrapulse'),
    ('LUMENIS', 'Stellar', 'lumenis', 'stellar'),
    ('LUMENIS', 'Splendor X', 'lumenis', 'splendor'),
    ('LUMENIS', 'NuEra Tight', 'lumenis', 'nuera'),
    ('CONTOURLINE', 'HiPro', 'contourline', 'hipro'),
    ('CONTOURLINE', 'MultiShape', 'contourline', 'multishape'),
    ('CONTOURLINE', 'Hive Pro', 'contourline', 'hivepro'),
    ('CONTOURLINE', 'FocuSkin', 'contourline', 'focuskin'),
    ('CONTOURLINE', 'UltraLift', 'contourline', 'ultralift'),
    ('CONTOURLINE', 'Reverso', 'contourline', 'reverso'),
    ('CONTOURLINE MED', 'HiPro Med', 'contourline_med', 'hipromed'),
    ('CONTOURLINE MED', 'Supreme Pro', 'contourline_med', 'supremepro'),
    ('BODY HEALTH', 'Crystal 3D', 'bodyhealth', 'crystal'),
    ('BODY HEALTH', 'Unyque Pro', 'bodyhealth', 'unyque'),
    ('BODY HEALTH', 'Enygma', 'bodyhealth', 'enygma'),
    ('BODY HEALTH', 'Iconyc', 'bodyhealth', 'iconyc'),
    ('VISBODY', 'S30', 'visbody', 's30'),
    ('VISBODY', 'M30', 'visbody', 'm30'),
    ('VISBODY', 'Creator 600', 'visbody', 'creator600'),
]

row = 2
for marca, equip, brand, term in equipamentos:
    url_meta = base + '?utm_brand=' + brand + '&utm_source=facebook&utm_medium=cpc&utm_term=' + term + '&utm_campaign={{campaign.name}}&utm_content={{ad.name}}'
    url_simples = base + '?utm_brand=' + brand + '&utm_source=facebook&utm_medium=cpc&utm_term=' + term

    ws1.cell(row=row, column=1, value=marca).font = brand_font
    ws1.cell(row=row, column=1).fill = brand_fills.get(marca, PatternFill('solid', fgColor='333333'))
    ws1.cell(row=row, column=1).alignment = Alignment(vertical='center')
    ws1.cell(row=row, column=2, value=equip).font = data_font
    ws1.cell(row=row, column=3, value=brand).font = data_font
    ws1.cell(row=row, column=4, value=term).font = data_font
    ws1.cell(row=row, column=5, value=url_meta).font = Font(name='Arial', size=9)
    ws1.cell(row=row, column=6, value=url_simples).font = link_font
    ws1.cell(row=row, column=6).hyperlink = url_simples

    for col in range(1, 7):
        ws1.cell(row=row, column=col).border = thin_border
    row += 1

# === ABA 2: Parametros UTM ===
ws2 = wb.create_sheet("Parametros UTM")

headers2 = ['Parametro', 'O que faz', 'Valores possiveis', 'Exemplo', 'Obrigatorio?']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 15

params = [
    ('utm_brand', 'Define a marca (logo, cores, perfis, glow)', 'contourline, contourline_med, lumenis, bodyhealth, visbody', 'lumenis', 'SIM'),
    ('utm_source', 'Identifica a plataforma de origem', 'facebook, instagram, google', 'facebook', 'SIM'),
    ('utm_medium', 'Tipo de midia paga', 'cpc, cpm, social, email', 'cpc', 'SIM'),
    ('utm_term', 'Identifica o equipamento', 'trilift, ultrapulse, stellar, hipro, s30, etc', 'trilift', 'SIM'),
    ('utm_campaign', 'Nome da campanha (dinamico do Meta)', '{{campaign.name}}', 'trilift_medicos_sp', 'RECOMENDADO'),
    ('utm_content', 'Variacao do criativo', '{{ad.name}}', 'video_demo_v2', 'OPCIONAL'),
]

for i, (param, desc, valores, exemplo, obrig) in enumerate(params, 2):
    ws2.cell(row=i, column=1, value=param).font = Font(name='Arial', bold=True, size=10, color='0563C1')
    ws2.cell(row=i, column=2, value=desc).font = data_font
    ws2.cell(row=i, column=3, value=valores).font = data_font
    ws2.cell(row=i, column=4, value=exemplo).font = data_font
    if obrig == 'SIM':
        fill_color = '90EE90'
    elif obrig == 'RECOMENDADO':
        fill_color = 'FFFFE0'
    else:
        fill_color = 'F0F0F0'
    ws2.cell(row=i, column=5, value=obrig).font = Font(name='Arial', bold=True, size=10)
    ws2.cell(row=i, column=5).fill = PatternFill('solid', fgColor=fill_color)
    ws2.cell(row=i, column=5).alignment = Alignment(horizontal='center')
    for col in range(1, 6):
        ws2.cell(row=i, column=col).border = thin_border
        ws2.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)

# === ABA 3: Instrucoes Meta Ads ===
ws3 = wb.create_sheet("Instrucoes Meta Ads")
ws3.column_dimensions['A'].width = 85

title_font = Font(name='Arial', bold=True, size=14, color='1a3a6b')
step_font = Font(name='Arial', bold=True, size=11)
text_font = Font(name='Arial', size=10)

instrucoes = [
    (title_font, 'Como usar as URLs no Gerenciador de Anuncios do Meta'),
    (text_font, ''),
    (step_font, 'PASSO 1: Criar campanha normalmente'),
    (text_font, 'Crie sua campanha no Meta Ads como de costume (objetivo, publico, orcamento).'),
    (text_font, ''),
    (step_font, 'PASSO 2: No nivel do ANUNCIO, encontrar o campo "URL do site"'),
    (text_font, 'Na secao "Destino", cole a URL da aba "URLs por Equipamento" correspondente.'),
    (text_font, ''),
    (step_font, 'PASSO 3: Usar a URL com parametros dinamicos'),
    (text_font, 'Cole a URL da coluna "URL Meta Ads (com dinamicos)".'),
    (text_font, 'O Meta substituira automaticamente {{campaign.name}} e {{ad.name}} pelos nomes reais.'),
    (text_font, ''),
    (step_font, 'PASSO 4: Verificar o formulario'),
    (text_font, 'Abra a URL da coluna "URL Simples (teste)" no navegador para verificar se carrega certo.'),
    (text_font, 'Confira: logo correto, equipamento no titulo, particulas na cor da marca.'),
    (text_font, ''),
    (step_font, 'PASSO 5: Para Instagram Ads'),
    (text_font, 'Troque utm_source=facebook por utm_source=instagram na URL.'),
    (text_font, 'O formulario identifica e grava a origem correta no RD Station.'),
    (text_font, ''),
    (step_font, 'PASSO 6: Para Google Ads'),
    (text_font, 'Troque utm_source=facebook por utm_source=google.'),
    (text_font, 'Use {campaign} no lugar de {{campaign.name}} (sintaxe do Google).'),
    (text_font, ''),
    (step_font, 'DICA: Parametros dinamicos do Meta Ads'),
    (text_font, '{{campaign.name}} = Nome da campanha'),
    (text_font, '{{adset.name}} = Nome do conjunto de anuncios'),
    (text_font, '{{ad.name}} = Nome do anuncio'),
    (text_font, '{{campaign.id}} = ID da campanha'),
    (text_font, '{{placement}} = Posicionamento (feed, stories, reels)'),
]

for i, (font, text) in enumerate(instrucoes, 1):
    c = ws3.cell(row=i, column=1, value=text)
    c.font = font
    c.alignment = Alignment(wrap_text=True)

import os
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'URLs_Campanhas_Meta_Ads.xlsx')
wb.save(output)
print("Saved: " + output)
